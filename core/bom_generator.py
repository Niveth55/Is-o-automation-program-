"""
BOM (Bill of Materials) Generator for ISO Piping.
Produces consolidated and per-line BOMs as CSV / Excel outputs.
"""
import os
import csv
from typing import List, Optional
from datetime import date

from core.models import PipeLine, PipeComponent
from config.settings import REPORTS_DIR


class BOMGenerator:
    """Generates Bill of Materials from processed piping lines."""

    def __init__(self, project_name: str = "Project", output_dir: str = REPORTS_DIR):
        self.project_name = project_name
        self.output_dir = output_dir
        os.makedirs(self.output_dir, exist_ok=True)

    # ------------------------------------------------------------------ #
    # Consolidate components across all lines
    # ------------------------------------------------------------------ #
    def consolidate(self, lines: List[PipeLine]) -> List[dict]:
        """
        Merge identical components (same description + size + material +
        schedule + pressure_class) and sum quantities / weights.
        """
        merged: dict = {}
        for line in lines:
            for comp in line.components:
                key = (
                    comp.description,
                    comp.size,
                    comp.size_2 or "",
                    comp.material,
                    comp.schedule,
                    comp.pressure_class,
                    comp.unit,
                    comp.end_connection,
                )
                if key not in merged:
                    merged[key] = {
                        "Description": comp.description,
                        "Size (inch)": comp.size if not comp.size_2
                                        else f"{comp.size} x {comp.size_2}",
                        "Material": comp.material,
                        "Schedule": comp.schedule,
                        "Pressure Class": comp.pressure_class,
                        "Unit": comp.unit,
                        "End Connection": comp.end_connection,
                        "Quantity": 0.0,
                        "Weight (kg)": 0.0,
                        "Lines": set(),
                    }
                merged[key]["Quantity"] += comp.quantity
                merged[key]["Weight (kg)"] += comp.weight_kg
                merged[key]["Lines"].add(line.line_no)

        result = []
        for i, (_, row) in enumerate(merged.items(), start=1):
            row = dict(row)
            row["Item No."] = i
            row["Lines"] = ", ".join(sorted(row["Lines"]))
            row["Weight (kg)"] = round(row["Weight (kg)"], 2)
            result.append(row)

        return result

    # ------------------------------------------------------------------ #
    # Per-line BOM
    # ------------------------------------------------------------------ #
    def line_bom(self, line: PipeLine) -> List[dict]:
        rows = []
        for comp in line.components:
            row = comp.to_dict()
            rows.append(row)
        return rows

    # ------------------------------------------------------------------ #
    # CSV export
    # ------------------------------------------------------------------ #
    def export_consolidated_csv(self, lines: List[PipeLine],
                                 filename: Optional[str] = None) -> str:
        rows = self.consolidate(lines)
        if not rows:
            return ""

        if filename is None:
            filename = f"{self.project_name}_BOM_Consolidated_{date.today()}.csv"
        path = os.path.join(self.output_dir, filename)

        fieldnames = [
            "Item No.", "Description", "Size (inch)", "Material",
            "Schedule", "Pressure Class", "End Connection",
            "Quantity", "Unit", "Weight (kg)", "Lines"
        ]
        with open(path, "w", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
            writer.writeheader()
            writer.writerows(rows)

        return path

    def export_line_csv(self, line: PipeLine,
                        filename: Optional[str] = None) -> str:
        rows = self.line_bom(line)
        if not rows:
            return ""

        safe_name = line.line_no.replace("/", "-").replace(" ", "_")
        if filename is None:
            filename = f"{self.project_name}_BOM_{safe_name}_{date.today()}.csv"
        path = os.path.join(self.output_dir, filename)

        fieldnames = [
            "Item No.", "Tag No.", "Description", "Size (inch)", "Material",
            "Schedule", "Pressure Class", "End Connection",
            "Quantity", "Unit", "Weight (kg)", "Remarks"
        ]
        with open(path, "w", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
            writer.writeheader()
            writer.writerows(rows)

        return path

    # ------------------------------------------------------------------ #
    # Excel export (uses openpyxl)
    # ------------------------------------------------------------------ #
    def export_excel(self, lines: List[PipeLine],
                     filename: Optional[str] = None) -> str:
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            print("openpyxl not installed — skipping Excel export.")
            return ""

        if filename is None:
            filename = f"{self.project_name}_BOM_{date.today()}.xlsx"
        path = os.path.join(self.output_dir, filename)

        wb = openpyxl.Workbook()

        # ----- Consolidated BOM sheet -----
        ws_cons = wb.active
        ws_cons.title = "Consolidated BOM"
        consolidated = self.consolidate(lines)
        self._write_excel_sheet(ws_cons, consolidated, [
            "Item No.", "Description", "Size (inch)", "Material",
            "Schedule", "Pressure Class", "End Connection",
            "Quantity", "Unit", "Weight (kg)", "Lines"
        ])

        # ----- Per-line BOM sheets -----
        for line in lines:
            safe = line.line_no[:30].replace("/", "-").replace(" ", "_")
            ws = wb.create_sheet(title=safe)
            # Header info rows
            ws.append([f"Line No.: {line.line_no}"])
            ws.append([f"Fluid: {line.fluid}   |   Spec: {line.pipe_spec}"])
            ws.append([f"Design: {line.design_pressure_bar} bar / {line.design_temp_c}°C"])
            ws.append([])
            rows = self.line_bom(line)
            self._write_excel_sheet(ws, rows, [
                "Item No.", "Tag No.", "Description", "Size (inch)", "Material",
                "Schedule", "Pressure Class", "End Connection",
                "Quantity", "Unit", "Weight (kg)", "Remarks"
            ], start_row=5)

        wb.save(path)
        return path

    def _write_excel_sheet(self, ws, rows: List[dict], columns: List[str],
                            start_row: int = 1):
        """Write header + data rows to an openpyxl worksheet."""
        try:
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            return

        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(bold=True, color="FFFFFF")
        header_align = Alignment(horizontal="center", vertical="center")

        # Write header
        ws.append(columns)
        header_row = ws[start_row]
        for cell in header_row:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align

        # Write data
        for row in rows:
            ws.append([row.get(col, "") for col in columns])

        # Auto-fit column widths
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    # ------------------------------------------------------------------ #
    # Summary report
    # ------------------------------------------------------------------ #
    def generate_summary_report(self, lines: List[PipeLine],
                                  filename: Optional[str] = None) -> str:
        if filename is None:
            filename = f"{self.project_name}_Summary_{date.today()}.txt"
        path = os.path.join(self.output_dir, filename)

        total_weight = sum(l.total_weight_kg for l in lines)
        total_comp = sum(l.component_count for l in lines)

        with open(path, "w") as f:
            f.write("=" * 60 + "\n")
            f.write(f"  ISO PIPING AUTOMATION — PROJECT SUMMARY REPORT\n")
            f.write(f"  Project: {self.project_name}\n")
            f.write(f"  Date: {date.today()}\n")
            f.write("=" * 60 + "\n\n")

            f.write(f"Total Piping Lines   : {len(lines)}\n")
            f.write(f"Total Components     : {total_comp}\n")
            f.write(f"Total Weight (kg)    : {round(total_weight, 2)}\n\n")

            f.write("-" * 60 + "\n")
            f.write(f"{'Line No.':<20} {'Fluid':<15} {'Spec':<10} "
                    f"{'Comp':>6} {'Weight (kg)':>12}\n")
            f.write("-" * 60 + "\n")
            for line in lines:
                f.write(
                    f"{line.line_no:<20} {line.fluid:<15} {line.pipe_spec:<10} "
                    f"{line.component_count:>6} {round(line.total_weight_kg, 2):>12}\n"
                )
            f.write("-" * 60 + "\n")
            f.write(f"{'TOTAL':<20} {'':<15} {'':<10} "
                    f"{total_comp:>6} {round(total_weight, 2):>12}\n")

        return path
